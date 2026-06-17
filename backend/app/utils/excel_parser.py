"""Excel parser for CLMS contract template files.

Parses the KHS Material Ketenagalistrikan sheet and Pemetaan Kontrak sheet
to extract clause objects with their variables and mandatory/optional status.
"""

import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from app.config import settings


class ClauseData:
    """Represents a single contract clause extracted from the Excel template."""

    def __init__(
        self,
        pasal_number: str,
        section_name: str,
        clause_text: str,
        variables: List[str],
        is_mandatory: bool = True,
        template_name: str = "KHS Material Ketenagalistrikan",
    ):
        self.pasal_number = pasal_number
        self.section_name = section_name
        self.clause_text = clause_text
        self.variables = variables
        self.is_mandatory = is_mandatory
        self.template_name = template_name

    def to_dict(self) -> Dict[str, Any]:
        """Convert clause data to dictionary."""
        return {
            "pasal_number": self.pasal_number,
            "section_name": self.section_name,
            "clause_text": self.clause_text,
            "variables": self.variables,
            "is_mandatory": self.is_mandatory,
            "template_name": self.template_name,
        }

    def __repr__(self) -> str:
        return (
            f"ClauseData(pasal='{self.pasal_number}', "
            f"section='{self.section_name}', "
            f"vars={len(self.variables)})"
        )


class ExcelParser:
    """Parser for CLMS Pemetaan Kontrak Excel files.

    Extracts contract clauses, variables, and mandatory/optional mappings
    from the structured Excel template.
    """

    def __init__(self, templates_dir: Optional[str] = None):
        """Initialize parser with templates directory path.

        Args:
            templates_dir: Path to the directory containing template files.
                          Defaults to the configured TEMPLATES_DIR.
        """
        if templates_dir:
            self.templates_dir = Path(templates_dir)
        else:
            self.templates_dir = settings.templates_absolute_path

    def __repr__(self) -> str:
        return f"ExcelParser(templates_dir='{self.templates_dir}')"

    def get_template_path(self, filename: str = "CLMS Pemetaan Kontrak.xlsx") -> Path:
        """Get the full path to a template file.

        Args:
            filename: Name of the template Excel file.

        Returns:
            Path object pointing to the template file.

        Raises:
            FileNotFoundError: If the template file does not exist.
        """
        path = self.templates_dir / filename
        if not path.exists():
            raise FileNotFoundError(f"Template file not found: {path}")
        return path

    def parse_khs_material_sheet(
        self, filepath: Optional[Path] = None
    ) -> List[ClauseData]:
        """Parse the KHS Material Ketenagalistrikan sheet.

        Extracts all clauses (Pasals) with their section names, full text,
        and associated variables.

        The sheet structure:
        - Column A: Pasal identifier (e.g., 'Pasal 1', 'Header')
        - Column B: Section/clause name
        - Column C: Full clause text with [PLACEHOLDER] variables
        - Column D: Variable/data field names

        Multi-row clauses: First row has Pasal in A, name in B, text in C,
        first variable in D. Subsequent rows have None in A/B/C with
        additional variables in D.

        Args:
            filepath: Optional path to the Excel file. Uses default if None.

        Returns:
            List of ClauseData objects representing all parsed clauses.
        """
        if filepath is None:
            filepath = self.get_template_path()

        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet_name = "KHS Material Ketenagalistrikan"

        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws = wb[sheet_name]
        clauses: List[ClauseData] = []

        # Track current clause being built
        current_pasal: Optional[str] = None
        current_section: Optional[str] = None
        current_text: Optional[str] = None
        current_variables: List[str] = []

        # Skip header row (row 1)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 4:
                continue

            col_a = str(row[0]).strip() if row[0] is not None else None
            col_b = str(row[1]).strip() if row[1] is not None else None
            col_c = str(row[2]).strip() if row[2] is not None else None
            col_d = str(row[3]).strip() if row[3] is not None else None

            # New clause starts when column A has a value
            if col_a:
                # Save previous clause if exists
                if current_pasal is not None:
                    clause = ClauseData(
                        pasal_number=current_pasal,
                        section_name=current_section or "",
                        clause_text=current_text or "",
                        variables=current_variables,
                        is_mandatory=True,
                    )
                    clauses.append(clause)

                # Start new clause
                current_pasal = col_a
                current_section = col_b or ""
                current_text = col_c or ""
                current_variables = []

                if col_d and col_d.lower() not in ("none", ""):
                    current_variables.append(col_d)
            else:
                # Continuation row - append variable from column D
                if col_d and col_d.lower() not in ("none", ""):
                    current_variables.append(col_d)

                # If there's text in column C, append to current clause text
                if col_c and col_c.lower() not in ("none", ""):
                    if current_text:
                        current_text += "\n" + col_c
                    else:
                        current_text = col_c

        # Don't forget the last clause
        if current_pasal is not None:
            clause = ClauseData(
                pasal_number=current_pasal,
                section_name=current_section or "",
                clause_text=current_text or "",
                variables=current_variables,
                is_mandatory=True,
            )
            clauses.append(clause)

        wb.close()
        return clauses

    def parse_pemetaan_kontrak_sheet(
        self, filepath: Optional[Path] = None
    ) -> Dict[str, Dict[str, str]]:
        """Parse the Pemetaan Kontrak sheet for clause status mapping.

        Maps clause names to their mandatory/optional status (M/O/P/X)
        for different contract types.

        The sheet structure:
        - Row 7: Contract type headers
        - Row 8: Sub-category headers
        - Rows 9-65: Clause names in col B, status codes in subsequent cols
        - KHS-specific columns: col 17 (Barang), col 18 (Jasa Lainnya),
          col 19 (Pekerjaan Konstruksi)

        Status codes:
        - M: Mandatory
        - O: Optional
        - P: Pilihan (Choice)
        - X: Not applicable

        Args:
            filepath: Optional path to the Excel file. Uses default if None.

        Returns:
            Dict mapping clause names to their status per contract sub-type.
        """
        if filepath is None:
            filepath = self.get_template_path()

        wb = load_workbook(filepath, read_only=True, data_only=True)
        # Note: sheet name has trailing space
        sheet_name = "Pemetaan Kontrak "

        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws = wb[sheet_name]
        status_mapping: Dict[str, Dict[str, str]] = {}

        # KHS sub-type columns (0-indexed in the tuple, so col 17 = index 16)
        khs_columns = {
            "Barang": 16,
            "Jasa Lainnya": 17,
            "Pekerjaan Konstruksi": 18,
        }

        # Parse rows 9 onwards for clause data
        for row in ws.iter_rows(min_row=9, max_row=65, values_only=True):
            if len(row) < 19:
                continue

            clause_name = str(row[1]).strip() if row[1] is not None else None
            if not clause_name or clause_name.lower() == "none":
                continue

            clause_status: Dict[str, str] = {}
            for sub_type, col_idx in khs_columns.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    status = str(row[col_idx]).strip().upper()
                    if status in ("M", "O", "P", "X"):
                        clause_status[sub_type] = status

            if clause_status:
                status_mapping[clause_name] = clause_status

        wb.close()
        return status_mapping

    def parse_all(
        self, filepath: Optional[Path] = None
    ) -> Dict[str, Any]:
        """Parse all sheets and return combined template data.

        Args:
            filepath: Optional path to the Excel file. Uses default if None.

        Returns:
            Dictionary containing clauses, status mapping, and template metadata.
        """
        if filepath is None:
            filepath = self.get_template_path()

        clauses = self.parse_khs_material_sheet(filepath)
        status_mapping = self.parse_pemetaan_kontrak_sheet(filepath)

        # Apply mandatory status from pemetaan to clauses
        for clause in clauses:
            if clause.section_name in status_mapping:
                statuses = status_mapping[clause.section_name]
                # If any sub-type marks it as mandatory, consider it mandatory
                clause.is_mandatory = any(
                    s == "M" for s in statuses.values()
                )

        return {
            "template_name": "KHS Material Ketenagalistrikan",
            "clauses": clauses,
            "status_mapping": status_mapping,
            "total_clauses": len(clauses),
            "total_variables": sum(len(c.variables) for c in clauses),
        }
